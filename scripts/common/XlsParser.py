# Contains utility classes for working with merged XLS localization files.

import os.path
import re
import sys
try:
	from openpyxl import *
	from openpyxl.styles import Font, Alignment
except:
	print "WARNING: Using XLS requires OpenPyXL installed!"

from LocParser import *

class XlsParser:

	DELIMITER = "\t"
	DASH = "\xC2\xAD" # the "other" dash character in UTF-8

	@staticmethod
	def parse(filename):		
		wb = load_workbook(filename)
		ws = wb.active
		
		currentRow = 1
		rowCount = ws.max_row
		
		locFiles = []
		locFile = None
		
		matches = []
		
		while currentRow < rowCount:			
			key = ws["A" + str(currentRow)].value
			value = ws["B" + str(currentRow)].value
			original = ws["C" + str(currentRow)].value
			comment = ws["D" + str(currentRow)].value			
			
			if key is None:
				key = ""
			if value is None:
				value = ""
			if original is None:
				original = ""
			if comment is None:
				comment = ""
				
			key = key.encode("ascii", "ignore")
			value = value.encode("ascii", "ignore")
			original = original.encode("ascii", "ignore")
			comment = comment.encode("ascii", "ignore")
			
			columns = [key, value, original, comment]
			
			matches.append(columns)
			
			currentRow = currentRow + 1
		
		for match in matches:
			key, value, original, comment = match
			columns = [key, value, original, comment]
			for i in xrange(len(columns)):				
				columns[i] = columns[i].replace(XlsParser.DASH, "-")
			if columns[0] == "###":
				if locFile != None:
					locFiles.append(locFile)
				language = columns[2]
				if language == "###":
					language = ""
				locFile = LocFile(columns[3], language, []) # cannot make [] a default argument because there's a bug in Python
			elif columns[0] != "" and locFile != None:
				locFile.entries.append(LocEntry(columns[0], columns[1], columns[2], columns[3]))
		if locFile != None:
			locFiles.append(locFile)		
		return locFiles
	
	@staticmethod
	def generateFile(locFiles):
		wb = Workbook()
		ws = wb.active
		font = Font(bold = True)		
		ws["A1"].value = "Key"
		ws["A1"].font = font
		ws["B1"].value = "Translation"
		ws["B1"].font = font
		ws["C1"].value = "Original"
		ws["C1"].font = font
		ws["D1"].value = "Comment"
		ws["D1"].font = font
		currentRow = 1;
		maxWidth = 10
		for locFile in locFiles:
			currentRow+=2			
			language = locFile.language
			if language == "":
				language = "###"
				
			ws["A" + str(currentRow)].value = "###"
			ws["B" + str(currentRow)].value = "###"
			ws["C" + str(currentRow)].value = language
			ws["D" + str(currentRow)].value = locFile.filename		
		
			currentRow+=1
			for locEntry in locFile.entries:
				currentRow += 1
				key, value, original, comment = locEntry.key, locEntry.value, locEntry.original, locEntry.comment
				columns = [key, value, original, comment]
				for i in xrange(len(columns)):
					 #prevents MS Excel from seeing "-" as a formula indicator
					columns[i] = columns[i].replace("-", XlsParser.DASH)	
				
				ws["A" + str(currentRow)].alignment = Alignment(horizontal='general', vertical='top',wrap_text = True)
				ws["B" + str(currentRow)].alignment = Alignment(horizontal='general', vertical='top',wrap_text = True)
				ws["C" + str(currentRow)].alignment = Alignment(horizontal='general', vertical='top',wrap_text = True)
				ws["D" + str(currentRow)].alignment = Alignment(horizontal='general', vertical='top',wrap_text = True)
					
				ws["A" + str(currentRow)].value = columns[0]
				ws["B" + str(currentRow)].value = columns[1]
				ws["C" + str(currentRow)].value = columns[2]
				ws["D" + str(currentRow)].value = columns[3]	
				
				if maxWidth < len(columns[0]):
					maxWidth = len(columns[0])

		ws.column_dimensions["A"].width = maxWidth
		ws.column_dimensions["B"].width = 40
		ws.column_dimensions["C"].width = 40
		ws.column_dimensions["D"].width = 40
		return wb
		
	@staticmethod
	def _makeEntry(key, translation, original, comment):
		return "\"%s\"%s\"%s\"%s\"%s\"%s\"%s\"\n" % (key.replace("\"", "\"\""), TsvParser.DELIMITER, translation.replace("\"", "\"\""),
			TsvParser.DELIMITER, original.replace("\"", "\"\""), TsvParser.DELIMITER, comment.replace("\"", "\"\""))
	
